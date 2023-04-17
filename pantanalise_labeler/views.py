from django.shortcuts import render
import openpyxl
xlsx_file = '../pantanalise_labeler/pantanalise_labeler/data/data.xlsx'

workbook = openpyxl.load_workbook(xlsx_file)
worksheet = workbook.active
max_rows = worksheet.max_row

def index(request):
    current_index = int(request.POST.get('current_index', '1'))
    current_row = worksheet.cell(row=current_index, column=1)
    current_text = current_row.offset(column=1).value
    current_feeling = current_row.offset(column=4).value

    if request.method == 'POST':
        if request.POST.get('next'):
            current_index += 1
            if current_index > max_rows:
                current_index = 1
            current_row = worksheet.cell(row=current_index, column=1)
            current_text = current_row.offset(column=1).value
            current_feeling = current_row.offset(column=4).value
            
        elif request.POST.get('prev'):
            current_index -= 1
            if current_index < 1:
                current_index = max_rows
            current_row = worksheet.cell(row=current_index, column=1)
            current_text = current_row.offset(column=1).value
            current_feeling = current_row.offset(column=4).value

        elif request.POST.get('neg'):
            current_row.offset(column=4).value = 'Neg'
            workbook.save(xlsx_file)
            current_feeling = 'Neg'
            
        elif request.POST.get('neu'):
            current_row.offset(column=4).value = 'Neu'
            workbook.save(xlsx_file)
            current_feeling = 'Neu'
            
        elif request.POST.get('pos'):
            current_row.offset(column=4).value = 'Pos'
            workbook.save(xlsx_file)
            current_feeling = 'Pos'

    return render(request, 'index.html', {'current_index': current_index, 'current_text': current_text, 'current_feeling': current_feeling, 'max_rows':max_rows})